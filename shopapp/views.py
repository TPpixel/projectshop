from io import BytesIO
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.conf import settings
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAuthenticated, SAFE_METHODS
from django.contrib.auth.models import User
from .models import Product, Category, Manufacturer, Basket, BasketItem, Order, OrderItem, Profile
from .serializers import (CategorySerializer, ManufacturerSerializer,
                          ProductSerializer, BasketSerializer,
                          BasketItemSerializer, OrderSerializer,
                          OrderItemSerializer, UserSerializer, ProfileSerializer)


def index(request):
    popular = Product.objects.order_by('?')[:6]
    categories = Category.objects.all()
    return render(request, 'shop/index.html', {
        'popular_products': popular,
        'categories': categories,
    })


def catalog(request):
    products = Product.objects.all()
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()

    category_id = request.GET.get('category')
    manufacturer_id = request.GET.get('manufacturer')
    query = request.GET.get('q')

    if category_id:
        products = products.filter(категория_id=category_id)
    if manufacturer_id:
        products = products.filter(производитель_id=manufacturer_id)
    if query:
        products = products.filter(
            Q(название__icontains=query) | Q(описание__icontains=query)
        )

    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'shop/catalog.html', {
        'page_obj': page_obj,
        'categories': categories,
        'manufacturers': manufacturers,
    })


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'shop/product_detail.html', {'product': product})


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    basket, _ = Basket.objects.get_or_create(пользователь=request.user)
    item, created = BasketItem.objects.get_or_create(
        корзина=basket,
        товар=product,
        defaults={'количество': 1}
    )
    if not created and item.количество < product.количество_на_складе:
        item.количество += 1
        item.save()
    return redirect('cart_view')


@login_required
def update_cart(request, item_id):
    item = get_object_or_404(BasketItem, pk=item_id, корзина__пользователь=request.user)
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        if 0 < quantity <= item.товар.количество_на_складе:
            item.количество = quantity
            item.save()
    return redirect('cart_view')


@login_required
def remove_from_cart(request, item_id):
    item = get_object_or_404(BasketItem, pk=item_id, корзина__пользователь=request.user)
    if request.method == 'POST':
        item.delete()
    return redirect('cart_view')


@login_required
def cart_view(request):
    basket, _ = Basket.objects.get_or_create(пользователь=request.user)
    items = basket.basketitem_set.all()
    return render(request, 'shop/cart.html', {
        'basket': basket,
        'items': items,
        'total': basket.общая_стоимость(),
    })


@login_required
def checkout(request):
    basket = Basket.objects.filter(пользователь=request.user).first()
    if not basket or not basket.basketitem_set.exists():
        return redirect('cart_view')

    if request.method == 'POST':
        address = request.POST.get('address', '')
        if not address:
            return render(request, 'shop/checkout.html', {'error': 'Введите адрес доставки'})

        order = Order.objects.create(
            пользователь=request.user,
            адрес_доставки=address,
        )
        total = Decimal('0.00')
        for item in basket.basketitem_set.all():
            OrderItem.objects.create(
                заказ=order,
                товар=item.товар,
                название_товара=item.товар.название,
                цена=item.товар.цена,
                количество=item.количество,
            )
            total += item.товар.цена * item.количество
        order.общая_стоимость = total
        order.save()

        wb = Workbook()
        ws = wb.active
        ws.title = f"Чек заказ #{order.id}"
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Чек заказа #{order.id}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = 'Покупатель:'
        ws['B3'] = request.user.username
        ws['A4'] = 'Email:'
        ws['B4'] = request.user.email
        ws['A5'] = 'Адрес доставки:'
        ws['B5'] = address
        ws['A6'] = 'Дата:'
        ws['B6'] = order.дата_создания.strftime('%d.%m.%Y %H:%M')

        headers = ['№', 'Товар', 'Цена', 'Количество', 'Сумма']
        thin = Side(style='thin')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=h)
            cell.font = Font(bold=True)
            cell.border = Border(top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal='center')

        for i, pos in enumerate(order.позиции.all(), 1):
            row = 8 + i
            ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=pos.название_товара)
            ws.cell(row=row, column=3, value=float(pos.цена))
            ws.cell(row=row, column=4, value=pos.количество).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=float(pos.стоимость()))

        total_row = 8 + order.позиции.count() + 1
        ws.cell(row=total_row, column=4, value='Итого:').font = Font(bold=True)
        ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=total_row, column=5, value=float(total)).font = Font(bold=True)

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        email = EmailMessage(
            subject=f'Чек заказа #{order.id} — {settings.EMAIL_HOST_USER}',
            body=f'Спасибо за покупку!\n\nНомер заказа: #{order.id}\nАдрес доставки: {address}\nОбщая стоимость: {total} BYN\n\nЧек прикреплён к письму.',
            from_email=settings.EMAIL_HOST_USER,
            to=[request.user.email],
        )
        email.attach(f'check_{order.id}.xlsx', output.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        try:
            email.send(fail_silently=False)
        except Exception:
            messages.warning(request, 'Заказ создан, но не удалось отправить письмо (проверьте настройки email).')

        basket.basketitem_set.all().delete()

        messages.success(request, f'Заказ #{order.id} оформлен! Чек отправлен на {request.user.email}.')
        return redirect('cart_view')

    return render(request, 'shop/checkout.html')


class IsAdminOrReadOnly(IsAuthenticatedOrReadOnly):
    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:
            return True
        if not request.user.is_authenticated:
            return False
        role = getattr(getattr(request.user, 'profile', None), 'роль', None)
        return request.user.is_staff or role in ('ADMIN', 'MANAGER')


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAdminOrReadOnly]


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]


class BasketViewSet(viewsets.ModelViewSet):
    queryset = Basket.objects.all()
    serializer_class = BasketSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        role = getattr(getattr(self.request.user, 'profile', None), 'роль', None)
        if self.request.user.is_staff or role in ('ADMIN', 'MANAGER'):
            return qs
        return qs.filter(пользователь=self.request.user)


class BasketItemViewSet(viewsets.ModelViewSet):
    queryset = BasketItem.objects.all()
    serializer_class = BasketItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        role = getattr(getattr(self.request.user, 'profile', None), 'роль', None)
        if self.request.user.is_staff or role in ('ADMIN', 'MANAGER'):
            return qs
        return qs.filter(корзина__пользователь=self.request.user)


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        role = getattr(getattr(self.request.user, 'profile', None), 'роль', None)
        if self.request.user.is_staff or role in ('ADMIN', 'MANAGER'):
            return qs
        return qs.filter(пользователь=self.request.user)


class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.all()
    serializer_class = OrderItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        role = getattr(getattr(self.request.user, 'profile', None), 'роль', None)
        if self.request.user.is_staff or role in ('ADMIN', 'MANAGER'):
            return qs
        return qs.filter(заказ__пользователь=self.request.user)


class ProfileViewSet(viewsets.ModelViewSet):
    queryset = Profile.objects.all()
    serializer_class = ProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        role = getattr(getattr(self.request.user, 'profile', None), 'роль', None)
        if self.request.user.is_staff or role in ('ADMIN', 'MANAGER'):
            return qs
        return qs.filter(пользователь=self.request.user)


@api_view(['GET', 'PATCH'])
def me(request):
    if not request.user.is_authenticated:
        return Response({'detail': 'Не авторизован'}, status=status.HTTP_401_UNAUTHORIZED)
    user = request.user
    if request.method == 'GET':
        serializer = UserSerializer(user)
        return Response(serializer.data)
    serializer = UserSerializer(user, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        if password != password2:
            return render(request, 'registration/register.html', {'error': 'Пароли не совпадают'})
        if User.objects.filter(username=username).exists():
            return render(request, 'registration/register.html', {'error': 'Пользователь уже существует'})
        user = User.objects.create_user(username=username, email=email, password=password)
        Profile.objects.create(пользователь=user)
        from django.contrib.auth import login
        login(request, user)
        return redirect('index')
    return render(request, 'registration/register.html')


@login_required
def profile(request):
    orders = Order.objects.filter(пользователь=request.user).order_by('-дата_создания')
    return render(request, 'shop/profile.html', {
        'orders': orders,
    })
